from selenium import webdriver
from openpyxl import load_workbook
from openpyxl import Workbook

# Load the source workbook
source_workbook = load_workbook('input.xlsx')
source_sheet = source_workbook.active

# Create a new workbook for the output
output_workbook = Workbook()
output_sheet = output_workbook.active
output_sheet.append(['Location', 'Review Value'])

# Create a new instance of the Chrome driver
driver = webdriver.Chrome()

# Navigate to the website
base_url = "https://www.cruiseline.com"

# Iterate over each location in column A of the source workbook
for row in source_sheet.iter_rows(min_row=2, values_only=True):
    location = row[0]

    # Navigate to the website
    driver.get(base_url)

    # Find the search input element by its ID
    search_input = driver.find_element("id", "cruiseline-search-mobile")

    # Set the value of the search input using JavaScript
    driver.execute_script("arguments[0].value = arguments[1];", search_input, location)

    # Submit the search form
    search_input.submit()

    # Wait for the page to load and retrieve the review element
    review_element = driver.find_element("class name", "rating-score")

    # Extract the value of the review from the element
    review_value = review_element.text.strip()

    # Append the location and review value to the output sheet
    output_sheet.append([location, review_value])

# Save the output workbook
output_workbook.save('bylocal.xlsx')

# Close the browser
driver.quit()
